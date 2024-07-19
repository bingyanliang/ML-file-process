import openpyxl
from openpyxl.styles import Border, Side, PatternFill, NamedStyle
from datetime import datetime

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def excel_date_to_datetime(excel_date):
    """Converts an Excel date to a datetime object."""
    return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + excel_date - 2)

def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    headers = data[0]
    data = data[1:]
    return headers, data, wb, ws

def write_excel(file_path, headers, data, wb, ws):
    posting_date_col_index = headers.index('Posting Date') + 1 if 'Posting Date' in headers else None
    trans_date_col_index = headers.index('Trans. Date') + 1 if 'Trans. Date' in headers else None

    # Clear existing data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Write new data
    for row_index, row_data in enumerate(data, start=2):
        for col_index, header in enumerate(headers, start=1):
            cell_value = row_data.get(header, '')
            # Convert Excel serial date to datetime if necessary
            if col_index == posting_date_col_index or col_index == trans_date_col_index:
                if isinstance(cell_value, (int, float)):  # Excel date serial number
                    cell_value = excel_date_to_datetime(int(cell_value))
            cell = ws.cell(row=row_index, column=col_index, value=cell_value)
            cell.border = thin_border
            if col_index == posting_date_col_index or col_index == trans_date_col_index:
                cell.number_format = 'MM/DD/YYYY'
        ws.cell(row=1, column=col_index).border = thin_border

    wb.save(file_path)

def update_category_col(headers, data, wb, ws):
    if 'Expense Category' not in headers:
        raise ValueError("The Excel file does not contain an 'Expense Category' column.")
    
    if 'Merchant Category' not in headers:
        raise ValueError("The Excel file does not contain an 'Merchant Category' column.")
    
    if 'Category' not in headers:
        headers.append('Category')
        category_col_idx = len(headers)
        ws.cell(row=1, column=category_col_idx, value='Category')
    else:
        category_col_idx = headers.index('Category') + 1

    expense_category_index = headers.index('Expense Category')
    merchant_category_index = headers.index('Merchant Category')
    category_index = headers.index('Category')

    for row_index, row_data in enumerate(data, start=2):
        expense_category = row_data[expense_category_index]
        merchant_category = row_data[merchant_category_index]
        category_value = ''
        if merchant_category in ['COURIER SERVICES-AIR OR GROUND, FREIGHT FORWA']:
            category_value = 'Shipping'
        elif expense_category in ['Travel', 'Transportation', 'Airlines']:
            category_value = 'Traveling'
        elif expense_category == 'Amusement and Entertainment':
            category_value = 'Meals Expense'
        elif expense_category == 'Utilities':
            category_value = 'Utilities'
        elif expense_category in ['Professional Services & Membership Organizations', 'Contracted Services', 'Business Services']:
            category_value = 'Office Expense'
        cell = ws.cell(row=row_index, column=category_index + 1, value=category_value)
        cell.border = thin_border

    ws.cell(row=1, column=category_col_idx).border = thin_border

def apply_row_colors(headers, data, wb, ws):
    matched_index = headers.index('Matched')

    good_style = NamedStyle(name="good_style")
    good_style.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    good_style.font = openpyxl.styles.Font(color="006100")
    good_style.border = thin_border

    neutral_style = NamedStyle(name="neutral_style")
    neutral_style.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    neutral_style.font = openpyxl.styles.Font(color="9C5700")
    neutral_style.border = thin_border

    if "good_style" not in wb.named_styles:
        wb.add_named_style(good_style)
    if "neutral_style" not in wb.named_styles:
        wb.add_named_style(neutral_style)

    for row_index, row_data in enumerate(data, start=2):
        matched_value = row_data[matched_index]
        style = good_style if matched_value == 'Y' else neutral_style

        for col_index in range(1, len(headers) + 1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.style = style