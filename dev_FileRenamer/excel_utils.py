import openpyxl
from openpyxl.styles import Border, Side

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    headers = data[0]
    data = data[1:]
    return headers, data, wb, ws

def write_excel(file_path, headers, data, wb, ws):

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    for row_index, row_data in enumerate(data, start=2):
        for col_index, header in enumerate(headers, start=1):
            cell=ws.cell(row=row_index, column=col_index, value=row_data.get(header, ''))
            cell.border = thin_border
            
    wb.save(file_path)

def update_category_col(headers, data, wb, ws):
    if 'Expense Category' not in headers:
        raise ValueError("The Excel file does not contain an 'Expense Category' column.")
    
    if 'Merchant Category' not in headers:
        raise ValueError("The Excel file does not contain an 'Merchant Category' column.")
    
    if 'Category' not in headers:
        headers.append('Category')
        category_col_idx = len(headers)
        ws.cell(row=1, column=category_col_idx, value='Category')
    else:
        category_col_idx = headers.index('Category') + 1

    expense_category_index = headers.index('Expense Category')
    merchant_category_index = headers.index('Merchant Category')
    category_index = headers.index('Category')

    for row_index, row_data in enumerate(data, start=2):
        expense_category = row_data[expense_category_index]
        merchant_category = row_data[merchant_category_index]
        category_value = ''
        if merchant_category in ['COURIER SERVICES-AIR OR GROUND, FREIGHT FORWA']:
            category_value = 'Shipping'
        elif expense_category in ['Travel', 'Transportation', 'Airlines']:
            category_value = 'Traveling'
        elif expense_category == 'Amusement and Entertainment':
            category_value = 'Meals Expense'
        elif expense_category == 'Utilities':
            category_value = 'Utilities'
        elif expense_category in ['Professional Services & Membership Organizations', 'Contracted Services', 'Business Services']:
            category_value = 'Office Expense'
        cell = ws.cell(row=row_index, column=category_index + 1, value=category_value)
        cell.border = thin_border
    
    ws.cell(row=1, column=category_col_idx).border = thin_border




